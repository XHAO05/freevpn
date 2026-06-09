import pandas as pd
import qrcode
import re
import os
import shutil
import base64
import json
import requests
import socket
import time
from datetime import datetime

from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def get_real_ip(host):
    try:
        return socket.gethostbyname(host)
    except Exception:
        return None

def get_ip_location(ip):
    if not ip or ip == "未知":
        return "未知"
    try:
        url = f"http://ip-api.com/json/{ip}?lang=zh-CN"
        response = requests.get(url, timeout=5).json()
        if response.get('status') == 'success':
            country = response.get('country', '')
            city = response.get('city', '')
            return f"{country} {city}".strip()
        else:
            return "查询失败"
    except Exception:
        return "网络超时"

def parse_node_info(link):
    try:
        protocol = link.split('://')[0].lower()
        host = "未知"
        
        if protocol == 'vmess':
            base64_str = link[8:]
            base64_str += "=" * ((4 - len(base64_str) % 4) % 4)
            decoded_str = base64.b64decode(base64_str).decode('utf-8')
            node_data = json.loads(decoded_str)
            host = node_data.get('add', '未知')
        else:
            match = re.search(r'@([^:/]+)', link)
            if match:
                host = match.group(1)
                
        ip = get_real_ip(host) if host != "未知" else "未知"
        return protocol, ip
        
    except Exception as e:
        return "未知", "未知"


def process_nodes(target_file):
    print(f"正在读取原文件 [{target_file}] 的第一个工作表...")
    
    try:
        df = pd.read_excel(target_file, sheet_name=0, header=None, usecols=[0])
        df.columns = ['原始链接']
        df = df.dropna(subset=['原始链接'])
    except Exception as e:
        print(f"❌ 读取 Excel 失败: {e}")
        return

    # --- 【WPS挤压数据自动拆解补丁】 ---
    if len(df) >= 1 and '\n' in str(df.iloc[0]['原始链接']):
        print("检测到数据被挤压在单个单元格，正在自动拆分换行...")
        raw_text = str(df.iloc[0]['原始链接'])
        split_links = [link.strip() for link in raw_text.split('\n') if link.strip()]
        df = pd.DataFrame({'原始链接': split_links})

    if df.empty:
        print("❌ 错误：表格里没有任何数据！请检查影刀粘贴动作。")
        return
    
    # --- 【清理防残留机制】彻底删除旧二维码重建 ---
    if os.path.exists("qr_codes"):
        shutil.rmtree("qr_codes")
    os.makedirs("qr_codes")
        
    print(f"开始处理 {len(df)} 个节点，正在查询 IP 并记录时间...")
    
    protocols, ips, locations, test_times, qr_paths = [], [], [], [], []    
    
    for i, row in df.iterrows():
        link = str(row['原始链接'])
        
        proto, ip = parse_node_info(link)
        location = get_ip_location(ip)
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 恢复你原本的超详细日志打印
        print(f"[{current_time}] 节点 {i+1} -> IP: {ip} -> 地区: {location}")
        
        time.sleep(1.2)
        
        qr_path = f"qr_codes/node_{i}.png"
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10, 
            border=2,
        )
        qr.add_data(link)
        qr.make(fit=True)
        img_raw = qr.make_image(fill_color="black", back_color="white")
        
        # 强行将原图拉伸到 1000x1000 像素高清底图
        img_high_res = img_raw.resize((1000, 1000), Image.NEAREST)
        img_high_res.save(qr_path)
        
        protocols.append(proto)
        ips.append(ip)
        locations.append(location)
        test_times.append(current_time)
        qr_paths.append(qr_path)
        
    df['协议类型'] = protocols
    df['服务器IP'] = ips
    df['IP归属地'] = locations
    df['测试时间'] = test_times
    df['二维码图'] = "" 
    
    if '序号' in df.columns:
        df = df.drop(columns=['序号'])
    df.insert(0, '序号', range(1, len(df) + 1))

    try:
        wb = load_workbook(target_file)
        
        if len(wb.sheetnames) > 1:
            target_sheet_name = wb.sheetnames[1]
            ws = wb[target_sheet_name]
        else:
            ws = wb.create_sheet("当天整理")
            target_sheet_name = ws.title

        # --- 【清理防残留机制】暴力删除目标表格中的所有旧行 ---
        ws._images = []
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        def set_col_width(col_name, width):
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width

        set_col_width('序号', 8)
        set_col_width('服务器IP', 16)
        set_col_width('IP归属地', 20)
        set_col_width('测试时间', 20)
        set_col_width('二维码图', 14)
        
        image_col_idx = df.columns.get_loc('二维码图') + 1
        image_col_letter = get_column_letter(image_col_idx)
        
        for i, path in enumerate(qr_paths):
            row_index = i + 2  
            ws.row_dimensions[row_index].height = 90
            
            img = OpenpyxlImage(path)
            img.width = 800
            img.height = 800
            ws.add_image(img, f"{image_col_letter}{row_index}")
            
        wb.save(target_file)
        wb.close()
        print(f"\n大功告成！已无损覆盖 [{target_file}]。请打开表格进行一键嵌入操作！")
        
    except PermissionError:
        print(f"\n⚠️ 写入失败：[{target_file}] 正在被打开！")
        print("💡 请先关闭该 Excel 表格，然后重新运行脚本。")

process_nodes("节点统计.xlsx")